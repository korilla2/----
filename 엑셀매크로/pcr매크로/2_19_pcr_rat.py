from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter.ttk as ttk
from tkinter import filedialog
from tkinter import *
from tkinter import messagebox

root = Tk()
root.title('PCR (RAT 양성) 입력기')
root.geometry('640x640')


def add_file():
    global files
    files = filedialog.askopenfilename(
        title='엑셀을 선택하세요', initialdir='C:/')

    list_file.insert(END, files)


def del_file():
    for index in reversed(list_file.curselection()):
        list_file.delete(index)


file_frame = Frame(root)
file_frame.pack(fill='x', padx=5, pady=5)

btn_add_file = Button(file_frame, padx=5, pady=5,
                      width=12, text='파일추가', command=add_file)
btn_add_file.pack(side='left')

btn_del_file = Button(file_frame, padx=5, pady=5,
                      width=12, text='선택삭제', command=del_file)
btn_del_file.pack(side='right')


list_frame = Frame(root)
list_frame.pack(fill='both', padx=5, pady=5)

scrollbar = Scrollbar(list_frame)
scrollbar.pack(side='right', fill='y')

list_file = Listbox(list_frame, selectmode='extended',
                    height=15, yscrollcommand=scrollbar.set)
list_file.pack(side='left', fill='both', expand=True)
scrollbar.config(command=list_file.yview)


frame_option = LabelFrame(root, text='옵션')
frame_option.pack(padx=5, pady=5)

frame_run = Frame(root)
frame_run.pack(fill='x')

btn_close = Button(frame_run, padx=5, pady=5, text='닫기',
                   width=12, command=root.quit)
btn_close.pack(side='right')

txt = Text(root, width=50, height=7)
txt.pack()
txt.insert(END, '이름을 입력하세요')


def btncmd():
    temp = txt.get('1.0', END)

    wb = load_workbook(files)

    ws = wb.active

    select = temp.strip('\n').split('\n')
    temp = []
    for i in range(1, ws.max_row+1):
        for j in select:
            if (ws.cell(row=i, column=4).value == j) or (ws.cell(row=i, column=4).value == j[0]+' '+j[1:]) or (ws.cell(row=i, column=4).value == j[0]+j[1]+' '+j[2:]):
                temp.append(j)
                ws[f'I{i}'] = 'PCR (RAT 양성)'

    for k in temp:
        if temp.count(k) > 1:
            messagebox.showinfo('중복 경고', f'중복이 발생하였습니다 {k}')
            y_color = PatternFill(start_color='ffff99',
                                  end_color='ffff99', fill_type='solid')

            for q in range(1, ws.max_row+1):
                if (ws.cell(row=q, column=4).value == k) or (ws.cell(row=q, column=4).value == k[0]+' '+k[1:]) or (ws.cell(row=q, column=4).value == k[0]+k[1]+' '+k[2:]):
                    ws.cell(row=q, column=4).fill = y_color
                    ws.cell(row=q, column=9).fill = y_color

            del temp[temp.index(k)]

    messagebox.showinfo('실행 완료', '종료하세요')

    wb.save(files)


def btncmd3():
    wb = load_workbook(files)
    ws = wb.active

    for i in range(1, ws.max_row+1):
        ws[f'I{i}'].value = 'PCR'

    for q in range(1, ws.max_row+1):
        y_color = PatternFill(start_color='ffffff',
                              end_color='ffffff', fill_type='solid')
        ws.cell(row=q, column=4).fill = y_color
        ws.cell(row=q, column=9).fill = y_color

    wb.save(files)


btn_txt = Button(root, text='RAT변환', command=btncmd, padx=5, pady=5)
btn_txt.pack(side='top')

btn_txt = Button(root, text='초기화', command=btncmd3, padx=5, pady=5)
btn_txt.pack(side='top')


root.resizable(False, False)
root.mainloop()
